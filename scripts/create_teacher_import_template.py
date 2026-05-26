#!/usr/bin/env python3
"""
Generate Excel template for Primary Teacher Import with Data Validation
St. Francis of Assisi School Portal
"""

import sys
import os

# Add path for openpyxl if available
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl not installed. Installing...")
    os.system("apt-get update -qq && apt-get install -y python3-openpyxl -qq")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

def create_primary_teacher_template():
    """Create Excel template with dropdowns for Primary Section teachers"""

    # Create workbook
    wb = Workbook()

    # Create main sheet
    ws = wb.active
    ws.title = "Primary Teachers"

    # Create reference data sheet (hidden)
    ref_sheet = wb.create_sheet("Reference Data")

    # Define column headers with descriptions
    headers = [
        ("name", "Full Name*", "Teacher's full name (Required)"),
        ("email", "Email Address*", "Unique email address (Required)"),
        ("phone", "Phone Number*", "Contact number with country code"),
        ("employee_id", "Employee ID*", "Unique staff ID (e.g., TEA001)"),
        ("username", "Username", "Login username (auto-generated if empty)"),
        ("password", "Password", "Login password (default: password123)"),
        ("status", "Status", "Account status (active/inactive/suspended)"),
        ("department", "Department*", "Primary or ECL only"),
        ("position", "Position", "Job title (e.g., Teacher, Senior Teacher)"),
        ("qualification", "Qualification", "Highest educational qualification"),
        ("specialization", "Specialization", "Leave EMPTY for Primary teachers"),
        ("join_date", "Joining Date", "Format: YYYY-MM-DD (e.g., 2024-01-15)"),
        ("basic_salary", "Basic Salary", "Monthly salary (e.g., 8500.00)"),
        ("nrc", "NRC", "National Registration Card number"),
        ("tpin", "TPIN", "Tax Identification Number"),
        ("account_number", "Bank Account", "Bank account number"),
        ("bank_name", "Bank Name", "Name of the bank"),
        ("bank_branch", "Bank Branch", "Bank branch location"),
        ("address", "Physical Address", "Complete physical address"),
        ("is_active", "Is Active?", "TRUE or FALSE"),
        ("is_class_teacher", "Class Teacher?", "TRUE if assigned to a specific class"),
        ("is_grade_teacher", "Grade Teacher?", "TRUE if responsible for entire grade"),
        ("grade", "Assigned Grade", "Grade name (required if Class/Grade teacher)"),
        ("class_section", "Class Section", "Section letter (A, B, C - required if Class teacher)"),
    ]

    # Style definitions
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    required_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Set column widths and headers
    for idx, (field, header, description) in enumerate(headers, 1):
        col_letter = get_column_letter(idx)
        ws.column_dimensions[col_letter].width = 20

        # Header cell
        cell = ws.cell(row=1, column=idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

        # Description/Help row
        desc_cell = ws.cell(row=2, column=idx)
        desc_cell.value = description
        desc_cell.font = Font(italic=True, size=9, color="666666")
        desc_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        desc_cell.border = border

        # Highlight required fields
        if "*" in header:
            for row in range(3, 103):  # Apply to 100 data rows
                ws.cell(row=row, column=idx).fill = required_fill

    # Set row heights
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 35

    # Freeze panes (freeze header and description rows)
    ws.freeze_panes = 'A3'

    # ==========================================
    # REFERENCE DATA (Hidden Sheet)
    # ==========================================

    # Status options
    ref_sheet['A1'] = 'Status'
    ref_sheet['A2'] = 'active'
    ref_sheet['A3'] = 'inactive'
    ref_sheet['A4'] = 'suspended'

    # Department options (Primary section only)
    ref_sheet['B1'] = 'Department'
    ref_sheet['B2'] = 'Primary'
    ref_sheet['B3'] = 'ECL'

    # Position options
    ref_sheet['C1'] = 'Position'
    ref_sheet['C2'] = 'Teacher'
    ref_sheet['C3'] = 'Senior Teacher'
    ref_sheet['C4'] = 'Head of Department'
    ref_sheet['C5'] = 'Deputy Head Teacher'

    # Qualification options
    ref_sheet['D1'] = 'Qualification'
    ref_sheet['D2'] = 'Certificate in Primary Education'
    ref_sheet['D3'] = 'Certificate in Early Childhood Education'
    ref_sheet['D4'] = 'Diploma in Primary Education'
    ref_sheet['D5'] = 'Diploma in Early Childhood Education'
    ref_sheet['D6'] = 'Bachelor of Education (Primary)'
    ref_sheet['D7'] = 'Bachelor of Arts in Education'
    ref_sheet['D8'] = 'Master of Education'
    ref_sheet['D9'] = 'Other'

    # Boolean options
    ref_sheet['E1'] = 'Boolean'
    ref_sheet['E2'] = 'TRUE'
    ref_sheet['E3'] = 'FALSE'

    # Primary Grades (ECL + Primary only)
    ref_sheet['F1'] = 'Primary Grades'
    ref_sheet['F2'] = 'Baby Class'
    ref_sheet['F3'] = 'Middle Class'
    ref_sheet['F4'] = 'Reception'
    ref_sheet['F5'] = 'Grade 1'
    ref_sheet['F6'] = 'Grade 2'
    ref_sheet['F7'] = 'Grade 3'
    ref_sheet['F8'] = 'Grade 4'
    ref_sheet['F9'] = 'Grade 5'
    ref_sheet['F10'] = 'Grade 6'
    ref_sheet['F11'] = 'Grade 7'

    # Class Sections
    ref_sheet['G1'] = 'Section'
    ref_sheet['G2'] = 'A'
    ref_sheet['G3'] = 'B'
    ref_sheet['G4'] = 'C'
    ref_sheet['G5'] = 'D'

    # Banks in Zambia
    ref_sheet['H1'] = 'Banks'
    ref_sheet['H2'] = 'Zanaco'
    ref_sheet['H3'] = 'Stanbic Bank'
    ref_sheet['H4'] = 'FNB Zambia'
    ref_sheet['H5'] = 'ABSA Bank'
    ref_sheet['H6'] = 'Standard Chartered'
    ref_sheet['H7'] = 'Bank of China'
    ref_sheet['H8'] = 'Indo Zambia Bank'
    ref_sheet['H9'] = 'Atlas Mara Bank'
    ref_sheet['H10'] = 'Access Bank'

    # ==========================================
    # DATA VALIDATION
    # ==========================================

    # Status dropdown (Column 7: status)
    dv_status = DataValidation(type="list", formula1="='Reference Data'!$A$2:$A$4", allow_blank=True)
    dv_status.error = 'Please select a valid status'
    dv_status.errorTitle = 'Invalid Status'
    ws.add_data_validation(dv_status)
    dv_status.add(f'G3:G102')

    # Department dropdown (Column 8: department)
    dv_dept = DataValidation(type="list", formula1="='Reference Data'!$B$2:$B$3", allow_blank=False)
    dv_dept.error = 'For Primary section, select Primary or ECL only'
    dv_dept.errorTitle = 'Invalid Department'
    ws.add_data_validation(dv_dept)
    dv_dept.add(f'H3:H102')

    # Position dropdown (Column 9: position)
    dv_position = DataValidation(type="list", formula1="='Reference Data'!$C$2:$C$5", allow_blank=True)
    dv_position.error = 'Please select a valid position'
    dv_position.errorTitle = 'Invalid Position'
    ws.add_data_validation(dv_position)
    dv_position.add(f'I3:I102')

    # Qualification dropdown (Column 10: qualification)
    dv_qual = DataValidation(type="list", formula1="='Reference Data'!$D$2:$D$9", allow_blank=True)
    dv_qual.error = 'Please select a valid qualification'
    dv_qual.errorTitle = 'Invalid Qualification'
    ws.add_data_validation(dv_qual)
    dv_qual.add(f'J3:J102')

    # Boolean dropdowns (is_active, is_class_teacher, is_grade_teacher)
    dv_bool = DataValidation(type="list", formula1="='Reference Data'!$E$2:$E$3", allow_blank=True)
    dv_bool.error = 'Please select TRUE or FALSE'
    dv_bool.errorTitle = 'Invalid Value'
    ws.add_data_validation(dv_bool)
    dv_bool.add(f'T3:T102')  # is_active (column 20)
    dv_bool.add(f'U3:U102')  # is_class_teacher (column 21)
    dv_bool.add(f'V3:V102')  # is_grade_teacher (column 22)

    # Grade dropdown (Column 23: grade)
    dv_grade = DataValidation(type="list", formula1="='Reference Data'!$F$2:$F$11", allow_blank=True)
    dv_grade.error = 'Please select a valid Primary or ECL grade'
    dv_grade.errorTitle = 'Invalid Grade'
    dv_grade.prompt = 'Select grade only if teacher is a Class Teacher or Grade Teacher'
    dv_grade.promptTitle = 'Grade Assignment'
    ws.add_data_validation(dv_grade)
    dv_grade.add(f'W3:W102')

    # Section dropdown (Column 24: class_section)
    dv_section = DataValidation(type="list", formula1="='Reference Data'!$G$2:$G$5", allow_blank=True)
    dv_section.error = 'Please select a valid section (A, B, C, D)'
    dv_section.errorTitle = 'Invalid Section'
    dv_section.prompt = 'Select section only if teacher is a Class Teacher'
    dv_section.promptTitle = 'Section Assignment'
    ws.add_data_validation(dv_section)
    dv_section.add(f'X3:X102')

    # Bank dropdown (Column 17: bank_name)
    dv_bank = DataValidation(type="list", formula1="='Reference Data'!$H$2:$H$10", allow_blank=True)
    dv_bank.error = 'Please select a valid bank'
    dv_bank.errorTitle = 'Invalid Bank'
    ws.add_data_validation(dv_bank)
    dv_bank.add(f'Q3:Q102')

    # ==========================================
    # ADD SAMPLE DATA
    # ==========================================

    sample_data = [
        [
            "John Mwale",
            "john.mwale@stfrancisofassisizm.com",
            "+260971234567",
            "TEA001",
            "john.mwale",
            "password123",
            "active",
            "Primary",
            "Teacher",
            "Diploma in Primary Education",
            "",  # No specialization for primary
            "2024-01-15",
            "8500.00",
            "123456/78/1",
            "1234567890",
            "0123456789",
            "Zanaco",
            "Cairo Road Branch",
            "Plot 123, Kabulonga, Lusaka",
            "TRUE",
            "TRUE",
            "FALSE",
            "Grade 5",
            "A"
        ],
        [
            "Mary Banda",
            "mary.banda@stfrancisofassisizm.com",
            "+260975111001",
            "TEA002",
            "mary.banda",
            "password123",
            "active",
            "ECL",
            "Teacher",
            "Certificate in Early Childhood Education",
            "",
            "2023-05-10",
            "7500.00",
            "234567/89/1",
            "2345678901",
            "1234567890",
            "Zanaco",
            "Woodlands Branch",
            "Plot 456, Chilenje, Lusaka",
            "TRUE",
            "TRUE",
            "FALSE",
            "Baby Class",
            "A"
        ],
        [
            "Peter Zulu",
            "peter.zulu@stfrancisofassisizm.com",
            "+260975222003",
            "TEA003",
            "peter.zulu",
            "password123",
            "active",
            "Primary",
            "Senior Teacher",
            "Bachelor of Education (Primary)",
            "",
            "2022-08-20",
            "10000.00",
            "345678/90/1",
            "3456789012",
            "2345678901",
            "Stanbic Bank",
            "Town Centre",
            "Plot 789, Matero, Lusaka",
            "TRUE",
            "FALSE",
            "TRUE",
            "Grade 7",
            ""
        ]
    ]

    # Add sample data starting from row 3
    for row_idx, row_data in enumerate(sample_data, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=False)

    # ==========================================
    # ADD INSTRUCTIONS SHEET
    # ==========================================

    instructions = wb.create_sheet("Instructions", 0)
    instructions.column_dimensions['A'].width = 100

    inst_data = [
        ["PRIMARY SECTION TEACHER IMPORT TEMPLATE", "header"],
        ["St. Francis of Assisi School Portal", "subheader"],
        ["", ""],
        ["IMPORTANT INSTRUCTIONS:", "section"],
        ["", ""],
        ["1. REQUIRED FIELDS (highlighted in yellow):", "bold"],
        ["   • Name, Email, Phone, Employee ID are mandatory", "text"],
        ["   • Each email and employee ID must be unique", "text"],
        ["", ""],
        ["2. PRIMARY SECTION ONLY:", "bold"],
        ["   • This template is for Primary and ECL teachers only", "text"],
        ["   • Department: Select 'Primary' or 'ECL' from dropdown", "text"],
        ["   • Specialization: Leave EMPTY (no subject specialization for primary teachers)", "text"],
        ["   • Valid Grades: Baby Class, Middle Class, Reception, Grade 1-7", "text"],
        ["", ""],
        ["3. DROPDOWNS:", "bold"],
        ["   • Status: active, inactive, or suspended", "text"],
        ["   • Department: Primary or ECL", "text"],
        ["   • Position: Teacher, Senior Teacher, HOD, Deputy Head Teacher", "text"],
        ["   • Qualification: Select from dropdown or type 'Other'", "text"],
        ["   • Is Active/Class Teacher/Grade Teacher: TRUE or FALSE", "text"],
        ["   • Grade: Select from dropdown (if Class or Grade teacher)", "text"],
        ["   • Section: A, B, C, or D (if Class teacher)", "text"],
        ["   • Bank Name: Select from list of Zambian banks", "text"],
        ["", ""],
        ["4. DATE FORMAT:", "bold"],
        ["   • Always use YYYY-MM-DD format (e.g., 2024-01-15)", "text"],
        ["", ""],
        ["5. CLASS TEACHER vs GRADE TEACHER:", "bold"],
        ["   • Class Teacher: Responsible for one specific class (e.g., Grade 5A)", "text"],
        ["     - Set is_class_teacher = TRUE", "text"],
        ["     - Specify both Grade AND Section", "text"],
        ["   • Grade Teacher: Responsible for entire grade level", "text"],
        ["     - Set is_grade_teacher = TRUE", "text"],
        ["     - Specify Grade only, leave Section empty", "text"],
        ["", ""],
        ["6. SAMPLE DATA:", "bold"],
        ["   • The template includes 3 sample teachers", "text"],
        ["   • You can modify or delete these and add your own", "text"],
        ["", ""],
        ["7. AFTER FILLING:", "bold"],
        ["   • Save the file as Excel (.xlsx)", "text"],
        ["   • Import through Admin Panel > Teachers > Import", "text"],
        ["", ""],
        ["For questions, contact the system administrator.", "text"],
    ]

    # Style instructions
    for row_idx, (text, style) in enumerate(inst_data, start=1):
        cell = instructions.cell(row=row_idx, column=1)
        cell.value = text
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        if style == "header":
            cell.font = Font(size=18, bold=True, color="366092")
            instructions.row_dimensions[row_idx].height = 25
        elif style == "subheader":
            cell.font = Font(size=14, color="666666")
        elif style == "section":
            cell.font = Font(size=13, bold=True, color="C65911")
        elif style == "bold":
            cell.font = Font(size=11, bold=True)
        else:
            cell.font = Font(size=11)

    # Hide reference data sheet
    ref_sheet.sheet_state = 'hidden'

    # Save workbook
    output_path = "/var/www/stfrancisofassisizm.com/projects/portal/public/downloads/templates/Primary_Teacher_Import_Template.xlsx"
    wb.save(output_path)
    print(f"✅ Template created successfully: {output_path}")
    return output_path

if __name__ == "__main__":
    try:
        path = create_primary_teacher_template()
        print(f"\n📊 Excel template with data validation created!")
        print(f"📁 Location: {path}")
        print(f"🌐 Download URL: http://portal.stfrancisofassisizm.com/downloads/templates/Primary_Teacher_Import_Template.xlsx")
    except Exception as e:
        print(f"❌ Error creating template: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
